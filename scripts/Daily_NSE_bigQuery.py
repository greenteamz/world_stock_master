import yfinance as yf
from datetime import datetime, timedelta
import gspread
from google.oauth2.service_account import Credentials
from google.cloud import bigquery
from google.api_core.exceptions import NotFound
from openpyxl import Workbook, load_workbook
import csv
import os
import pytz
import time
import logging

# Define IST timezone
IST = pytz.timezone('Asia/Kolkata')

# Current IST datetime
ist_now = datetime.now(IST)

# Extract the date part from IST datetime
ist_date = ist_now.date()

# Generate log and CSV file names 
log_filename = f"log_{ist_now.strftime('%Y-%m-%d_%H-%M-%S')}.txt"
master_log_filename = f"Log_Master_NSE_BigQuery_test_dump.txt"
csv_filename = f"NSE_Stock_Master_BQ_test_dump.csv"  # Append data for the same day
csv_filename_daily = f"NSE_Stock_{ist_now.strftime('%Y-%m-%d_%H-%M-%S')}_test_dump.csv"  # Append data for the same day
excel_filename = f"NSE_Stock_Master_DataLake_test_dump.xlsx"  # Excel file for today

# Paths for logs, CSV, and Excel
MASTER_LOG_FILE_PATH = os.path.join("logs", master_log_filename)
LOG_FILE_PATH = os.path.join("logs", log_filename)
MASTER_CSV_FILE_PATH = os.path.join("csv", csv_filename)
Daily_CSV_FILE_PATH = os.path.join("csv", csv_filename_daily)
EXCEL_FILE_PATH = os.path.join("excel", excel_filename)

# Ensure directories exist
os.makedirs("master_log", exist_ok=True)
os.makedirs("logs", exist_ok=True)
os.makedirs("csv", exist_ok=True)
os.makedirs("excel", exist_ok=True)

# Log function
def log_message(message):
    """Log messages to a file and print to console."""
    timestamp = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    with open(LOG_FILE_PATH, "a") as log_file:
        log_file.write(f"[{timestamp}] {message}\n")
    with open(MASTER_LOG_FILE_PATH, "a") as master_log_file:
        master_log_file.write(f"[{timestamp}] {message}\n")
    print(f"[{timestamp}] {message}")

# Authenticate using the same service_account.json for both BigQuery and Google Sheets
SERVICE_ACCOUNT_FILE = "service_account.json"

# Google Sheets authentication
gc = gspread.service_account(filename=SERVICE_ACCOUNT_FILE)


# Open Google Spreadsheet
spreadsheet = gc.open('NSE_symbol')  # Replace with your Google Sheet name
#source_worksheet = spreadsheet.worksheet('symbol')  # Replace with your sheet name
source_worksheet = spreadsheet.worksheet('symbol_test')  # Test sheet name

# Fetch all stock symbols from the first column
symbols = source_worksheet.col_values(1)[1:]  # Skip header row
symbols = [symbol if symbol.endswith('.NS') else f"{symbol}.NS" for symbol in symbols]

# Define BigQuery dataset and table with the project ID
PROJECT_ID = "stockautomation-442015"  # Replace with your project ID
BQ_DATASET = "nse_stock_test_dump"  # Replace with your dataset name
BQ_TABLE = f"{PROJECT_ID}.{BQ_DATASET}.daily_nse_stock_data_test"  # Fully-qualified table name

# Define schema for BigQuery table
headers = [
    "industry", "sector", "fullTimeEmployees", "auditRisk", "boardRisk",
    "compensationRisk", "shareHolderRightsRisk", "overallRisk", "maxAge", "priceHint",
    "previousClose", "open", "dayLow", "dayHigh", "regularMarketPreviousClose",
    "regularMarketOpen", "regularMarketDayLow", "regularMarketDayHigh", "dividendRate",
    "dividendYield", "exDividendDate", "payoutRatio", "fiveYearAvgDividendYield", "beta",
    "trailingPE", "forwardPE", "volume", "regularMarketVolume", "averageVolume",
    "averageVolume10days", "averageDailyVolume10Day", "marketCap", "fiftyTwoWeekLow",
    "fiftyTwoWeekHigh", "priceToSalesTrailing12Months", "fiftyDayAverage",
    "twoHundredDayAverage", "trailingAnnualDividendRate", "trailingAnnualDividendYield",
    "enterpriseValue", "profitMargins", "floatShares", "sharesOutstanding",
    "heldPercentInsiders", "heldPercentInstitutions", "impliedSharesOutstanding",
    "bookValue", "priceToBook", "earningsQuarterlyGrowth", "trailingEps", "forwardEps",
    "52WeekChange", "lastDividendValue", "lastDividendDate", "exchange", "quoteType",
    "symbol", "shortName", "longName", "currentPrice", "targetHighPrice", "targetLowPrice",
    "targetMeanPrice", "targetMedianPrice", "recommendationMean", "recommendationKey",
    "numberOfAnalystOpinions", "totalCash", "totalCashPerShare", "ebitda", "totalDebt",
    "quickRatio", "currentRatio", "totalRevenue", "debtToEquity", "revenuePerShare",
    "returnOnAssets", "returnOnEquity", "freeCashflow", "operatingCashflow",
    "earningsGrowth", "revenueGrowth", "grossMargins", "ebitdaMargins", "operatingMargins"
]

# Define a data type mapping for headers
data_type_map = {
    "industry": "STRING",
    "sector": "STRING",
    "fullTimeEmployees": "INTEGER",  # Integer field
    "auditRisk": "FLOAT",
    "boardRisk": "FLOAT",
    "compensationRisk": "FLOAT",
    "shareHolderRightsRisk": "FLOAT",
    "overallRisk": "FLOAT",
    "maxAge": "INTEGER",
    "priceHint": "INTEGER",
    "previousClose": "FLOAT",
    "open": "FLOAT",
    "dayLow": "FLOAT",
    "dayHigh": "FLOAT",
    "regularMarketPreviousClose": "FLOAT",
    "regularMarketOpen": "FLOAT",
    "regularMarketDayLow": "FLOAT",
    "regularMarketDayHigh": "FLOAT",
    "dividendRate": "FLOAT",
    "dividendYield": "FLOAT",
    "exDividendDate": "DATE",
    "payoutRatio": "FLOAT",
    "fiveYearAvgDividendYield": "FLOAT",
    "beta": "FLOAT",
    "trailingPE": "FLOAT",
    "forwardPE": "FLOAT",
    "volume": "INTEGER",
    "regularMarketVolume": "INTEGER",
    "averageVolume": "INTEGER",
    "averageVolume10days": "INTEGER",
    "averageDailyVolume10Day": "INTEGER",
    "marketCap": "INTEGER",
    "fiftyTwoWeekLow": "FLOAT",
    "fiftyTwoWeekHigh": "FLOAT",
    "priceToSalesTrailing12Months": "FLOAT",
    "fiftyDayAverage": "FLOAT",
    "twoHundredDayAverage": "FLOAT",
    "trailingAnnualDividendRate": "FLOAT",
    "trailingAnnualDividendYield": "FLOAT",
    "enterpriseValue": "INTEGER",
    "profitMargins": "FLOAT",
    "floatShares": "INTEGER",
    "sharesOutstanding": "INTEGER",
    "heldPercentInsiders": "FLOAT",
    "heldPercentInstitutions": "FLOAT",
    "impliedSharesOutstanding": "INTEGER",
    "bookValue": "FLOAT",
    "priceToBook": "FLOAT",
    "earningsQuarterlyGrowth": "FLOAT",
    "trailingEps": "FLOAT",
    "forwardEps": "FLOAT",
    "52WeekChange": "FLOAT",
    "lastDividendValue": "FLOAT",
    "lastDividendDate": "DATE",
    "exchange": "STRING",
    "quoteType": "STRING",
    "symbol": "STRING",
    "shortName": "STRING",
    "longName": "STRING",
    "currentPrice": "FLOAT",
    "targetHighPrice": "FLOAT",
    "targetLowPrice": "FLOAT",
    "targetMeanPrice": "FLOAT",
    "targetMedianPrice": "FLOAT",
    "recommendationMean": "FLOAT",
    "recommendationKey": "STRING",
    "numberOfAnalystOpinions": "INTEGER",
    "totalCash": "INTEGER",
    "totalCashPerShare": "FLOAT",
    "ebitda": "INTEGER",
    "totalDebt": "INTEGER",
    "quickRatio": "FLOAT",
    "currentRatio": "FLOAT",
    "totalRevenue": "INTEGER",
    "debtToEquity": "FLOAT",
    "revenuePerShare": "FLOAT",
    "returnOnAssets": "FLOAT",
    "returnOnEquity": "FLOAT",
    "freeCashflow": "INTEGER",
    "operatingCashflow": "INTEGER",
    "earningsGrowth": "FLOAT",
    "revenueGrowth": "FLOAT",
    "grossMargins": "FLOAT",
    "ebitdaMargins": "FLOAT",
    "operatingMargins": "FLOAT",
}

ROW_COUNTER_FILE = "master/row_counter.txt"

# Initialize row_insert_order
def initialize_row_counter():
    if not os.path.exists(ROW_COUNTER_FILE):
        with open(ROW_COUNTER_FILE, "w") as f:
            f.write("1")  # Start counter at 0

def get_current_row_counter():
    with open(ROW_COUNTER_FILE, "r") as f:
        return int(f.read().strip())

def update_row_counter(new_value):
    with open(ROW_COUNTER_FILE, "w") as f:
        f.write(str(new_value))

# Initialize the row counter if not already done
initialize_row_counter()


# Add "Previous Day Date" to headers
# PREVIOUS_DAY_DATE = (ist_date - timedelta(days=1)).strftime('%Y-%m-%d')
PREVIOUS_DAY_DATETIME = (ist_now - timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
headers_with_date = ["row_insert_order", "PreviousDayDate", "Symbol_Input"] + headers

def ensure_dataset_exists():
    try:
        bq_client.get_dataset(BQ_DATASET)
        log_message(f"Dataset '{BQ_DATASET}' exists.")
    except NotFound:
        dataset = bigquery.Dataset(f"{PROJECT_ID}.{BQ_DATASET}")
        bq_client.create_dataset(dataset)
        log_message(f"Created dataset '{BQ_DATASET}'.")

def ensure_table_exists():
    try:
        # Check if the table already exists
        table = bq_client.get_table(BQ_TABLE)
        log_message(f"Table '{BQ_TABLE}' already exists.")
    except NotFound:
        # Table does not exist, create it
        # Build the schema dynamically
        schema = [bigquery.SchemaField("row_insert_order", "INTEGER"), bigquery.SchemaField("PreviousDayDate", "DATETIME"), bigquery.SchemaField("Symbol_Input", "STRING"),] + [
                bigquery.SchemaField(header, data_type_map.get(header, "STRING"))
                for header in headers
                ]
        
        #schema = [bigquery.SchemaField("PreviousDayDate", "DATETIME"), bigquery.SchemaField("Symbol_Input", "STRING")] + [
        #    bigquery.SchemaField(header, "STRING") for header in headers
        #]

        table = bigquery.Table(BQ_TABLE, schema=schema)
        bq_client.create_table(table)
        log_message(f"Created table '{BQ_TABLE}'.")
    except Exception as e:
        log_message(f"Error ensuring table exists: {e}")

def append_to_csv(data_row):
    """Append a row of data to the CSV file, adding the header only if it's a new file."""
    write_header = not os.path.exists(MASTER_CSV_FILE_PATH)  # Check if file exists

    with open(MASTER_CSV_FILE_PATH, mode="a", newline="") as csv_file:
        writer = csv.writer(csv_file)
        if write_header:
            writer.writerow(["row_insert_order", "PreviousDayDate", "Symbol_Input"] + headers)  # Add header row
            log_message(f"Header added to CSV file: {MASTER_CSV_FILE_PATH}")
        writer.writerow(data_row)
        log_message(f"Appended data to CSV file: {MASTER_CSV_FILE_PATH}")

    """Append a row of data to the CSV file, adding the header only if it's a new file."""
    write_header = not os.path.exists(Daily_CSV_FILE_PATH)  # Check if file exists

    with open(Daily_CSV_FILE_PATH, mode="a", newline="") as csv_file:
        writer = csv.writer(csv_file)
        if write_header:
            writer.writerow(["row_insert_order", "PreviousDayDate", "Symbol_Input"] + headers)  # Add header row
            log_message(f"Header added to CSV file: {Daily_CSV_FILE_PATH}")
        writer.writerow(data_row)
        log_message(f"Appended data to CSV file: {Daily_CSV_FILE_PATH}")

def append_to_excel(data_row):
    """Append data to an Excel sheet, creating a new sheet for the day."""
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            workbook = load_workbook(EXCEL_FILE_PATH)
        else:
            workbook = Workbook()

        sheet_name = f"NSE_{ist_date}"
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
            sheet = workbook[sheet_name]
            sheet.append(["row_insert_order", "PreviousDayDate", "Symbol_Input"] + headers)  # Add header
        else:
            sheet = workbook[sheet_name]

        sheet.append(data_row)
        workbook.save(EXCEL_FILE_PATH)
        log_message(f"Data appended to Excel file: {EXCEL_FILE_PATH}")
    except Exception as e:
        log_message(f"Error saving to Excel: {e}")

def fetch_and_update_stock_data(symbol):
    try:
        log_message(f"Fetching data for {symbol}...")
        stock = yf.Ticker(symbol)
        info = stock.info

        # Read the current row counter
        current_counter = get_current_row_counter()
        
        # PREVIOUS_DAY_DATE = (ist_date - timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
        PREVIOUS_DAY_DATETIME = (ist_now - timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
        # Extract data and include the Previous Day Date
        info_row = [current_counter, PREVIOUS_DAY_DATETIME, symbol] + [info.get(key, '') for key in headers]

        # Increment row_insert_order for the next row
        current_counter += 1
        update_row_counter(current_counter)
        
        # Append data to CSV and Excel
        append_to_csv(info_row)
        append_to_excel(info_row)
        return info_row
    except Exception as e:
        log_message(f"Error fetching data for {symbol}: {e}")
        return None

# Add process data
def preprocess_data(csv_file_path):
    """
    Preprocess the CSV file to ensure data types are correct based on the BigQuery schema.
    If incorrect types are detected, log the error and attempt to fix them.
    """

    processed_rows = []
    errors = []

    try:
        with open(csv_file_path, "r") as csv_file:
            reader = csv.DictReader(csv_file)
            for row in reader:
                processed_row = {}
                for key, value in row.items():
                    expected_type = data_type_map.get(key, "STRING")
                    try:
                        if expected_type == "INTEGER":
                            processed_row[key] = int(value) if value else None
                        elif expected_type == "FLOAT":
                            processed_row[key] = float(value) if value else None
                        elif expected_type == "DATETIME":
                            processed_row[key] = (
                                datetime.strptime(value, "%Y-%m-%d %H:%M:%S") 
                                if value 
                                else datetime(1990, 1, 1, 0, 0, 0)  # Default to '1990-01-01 00:00:00' if no value is provided
                            )
                        elif expected_type == "DATE":
                            try:
                                # Check if value is a Unix timestamp and convert it
                                if value.isdigit():
                                    processed_row[key] = datetime.fromtimestamp(int(value)).date()
                                else:
                                    # Parse date string in the format "YYYY-MM-DD"
                                    processed_row[key] = datetime.strptime(value, "%Y-%m-%d").date()
                            except Exception:
                                # Handle invalid or missing date values with a default date
                                processed_row[key] = datetime(1990, 1, 1).date()

                        else:  # STRING
                            processed_row[key] = value
                    except ValueError as ve:
                        errors.append(
                            f"Field '{key}' with value '{value}' failed type conversion to {expected_type}: {ve}"
                        )
                        processed_row[key] = None  # Set to None on error
                processed_rows.append(processed_row)
    except Exception as e:
        log_message(f"Error reading or processing CSV file: {e}")
    
    # Log errors, if any
    if errors:
        log_message(f"Data type errors detected during preprocessing:\n" + "\n".join(errors))
    
    return processed_rows

def load_data_to_bigquery():
    """Load data from the preprocessed CSV file into BigQuery."""
    try:
        processed_data = preprocess_data(Daily_CSV_FILE_PATH)
        
        # Write processed data back to a temporary CSV for BigQuery loading
        temp_csv_path = "temp_processed.csv"
        with open(temp_csv_path, "w", newline="") as temp_csv:
            writer = csv.DictWriter(temp_csv, fieldnames=processed_data[0].keys())
            writer.writeheader()  # Write headers
            writer.writerows(processed_data)  # Write processed rows
        
        # Load the processed data into BigQuery
        with open(temp_csv_path, "rb") as csv_file:
            job_config = bigquery.LoadJobConfig(
                source_format=bigquery.SourceFormat.CSV,
                skip_leading_rows=1,  # Skip header row
                write_disposition="WRITE_APPEND",  # Append data schema=schema
                autodetect=False,
            )
            load_job = bq_client.load_table_from_file(
                csv_file, BQ_TABLE, job_config=job_config
            )
            load_job.result()  # Wait for the job to complete
            log_message(f"Data successfully loaded to BigQuery from {temp_csv_path}.")
    except Exception as e:
        log_message(f"Error loading data to BigQuery: {e}")

### Disable this Query #### 
'''
def load_data_to_bigquery():
    """Load data from the CSV file into BigQuery."""
    try:
        with open(CSV_FILE_PATH, "rb") as csv_file:
            job_config = bigquery.LoadJobConfig(
                source_format=bigquery.SourceFormat.CSV,
                skip_leading_rows=1,  # Skip header row
                write_disposition="WRITE_APPEND",  # Append data
                autodetect=True,
            )
            load_job = bq_client.load_table_from_file(
                csv_file, BQ_TABLE, job_config=job_config
            )
            load_job.result()  # Wait for the job to complete
            log_message(f"Data successfully loaded to BigQuery from {CSV_FILE_PATH}.")
    except Exception as e:
        log_message(f"Error loading data to BigQuery: {e}")
'''



# Process each symbol
for symbol in symbols:
    fetch_and_update_stock_data(symbol)

# BigQuery authentication
bq_client = bigquery.Client.from_service_account_json(SERVICE_ACCOUNT_FILE)

# Ensure dataset and table exist in BigQuery
ensure_dataset_exists()
ensure_table_exists()

# Load the data into BigQuery from the CSV file
load_data_to_bigquery()

log_message("Script execution completed.")
