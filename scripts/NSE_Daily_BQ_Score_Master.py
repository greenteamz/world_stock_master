import yfinance as yf
from datetime import datetime, timedelta
import gspread
from google.oauth2.service_account import Credentials
from google.cloud import bigquery
from google.api_core.exceptions import NotFound
from openpyxl import Workbook, load_workbook
import csv
import os
import pytz
import time
import logging
import pandas as pd
from collections import defaultdict

# Define IST timezone
IST = pytz.timezone('Asia/Kolkata')

# Current IST datetime
ist_now = datetime.now(IST)

# Extract the date part from IST datetime
ist_date = ist_now.date()

# Generate log and CSV file names 
log_filename = f"log_Daily_{ist_now.strftime('%Y-%m-%d_%H-%M-%S')}_score_master.txt"
master_log_filename = f"Log_Master_NSE_BigQuery_score_master.txt"
csv_filename = f"NSE_Stock_Master_BQ_score_master.csv"  # Append data for the same day
csv_filename_daily = f"NSE_Stock_Daily_{ist_now.strftime('%Y-%m-%d_%H-%M-%S')}_score_master.csv"  # Append data for the same day
excel_filename = f"NSE_Stock_Master_DataLake_score_master.xlsx"  # Excel file for today

# Define base directory
BASE_DIR = "NSE"

# Subdirectories under NSE
MASTER_DIR = os.path.join(BASE_DIR, "master_nse")
LOGS_DIR = os.path.join(BASE_DIR, "logs_nse")
CSV_DIR = os.path.join(BASE_DIR, "csv_nse")

# Paths for logs, CSV, and Excel
MASTER_LOG_FILE_PATH = os.path.join(MASTER_DIR, master_log_filename)
LOG_FILE_PATH = os.path.join(LOGS_DIR, log_filename)
MASTER_CSV_FILE_PATH = os.path.join(MASTER_DIR, csv_filename)
Daily_CSV_FILE_PATH  = os.path.join(CSV_DIR, csv_filename_daily)
EXCEL_FILE_PATH = os.path.join(MASTER_DIR, excel_filename)

# Ensure all required directories exist
os.makedirs(MASTER_DIR, exist_ok=True)
os.makedirs(LOGS_DIR, exist_ok=True)
os.makedirs(CSV_DIR, exist_ok=True)

# Log function
def log_message(message):
    """Log messages to a file and print to console."""
    timestamp = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    with open(LOG_FILE_PATH, "a") as log_file:
        log_file.write(f"[{timestamp}] {message}\n")
    with open(MASTER_LOG_FILE_PATH, "a") as master_log_file:
        master_log_file.write(f"[{timestamp}] {message}\n")
    print(f"[{timestamp}] {message}")

# Authenticate using the same service_account.json for both BigQuery and Google Sheets
SERVICE_ACCOUNT_FILE = "service_account.json"

# Google Sheets authentication
gc = gspread.service_account(filename=SERVICE_ACCOUNT_FILE)


# Open Google Spreadsheet
spreadsheet = gc.open('NSE_symbol')  # Replace with your Google Sheet name
#source_worksheet = spreadsheet.worksheet('symbol')  # Replace with your sheet name
source_worksheet = spreadsheet.worksheet('symbol')  # Test sheet name

# Fetch all stock symbols from the first column
symbols = source_worksheet.col_values(1)[1:]  # Skip header row
symbols = [symbol if symbol.endswith('.NS') else f"{symbol}.NS" for symbol in symbols]

# Define schema for BigQuery table
headers = [
    "fullTimeEmployees", "auditRisk", "boardRisk",
    "compensationRisk", "shareHolderRightsRisk", "overallRisk", "maxAge", "priceHint",
    "regularMarketOpen", "regularMarketDayLow", "regularMarketDayHigh", "dividendRate",
    "dividendYield", "exDividendDate", "payoutRatio", "fiveYearAvgDividendYield", "beta",
    "volume", "regularMarketVolume", "averageVolume",
    "returnOnAssets", "returnOnEquity", "freeCashflow", "operatingCashflow",
    "twoHundredDayAverage", "trailingAnnualDividendRate", "trailingAnnualDividendYield",
    "enterpriseValue", "floatShares", "sharesOutstanding",
    "heldPercentInsiders", "heldPercentInstitutions", "impliedSharesOutstanding",
    "bookValue", "priceToBook", "earningsQuarterlyGrowth", "trailingEps", "forwardEps",
    "52WeekChange", "lastDividendValue", "lastDividendDate", "exchange", "quoteType", 
    "totalCash", "totalCashPerShare", "ebitda", "totalDebt",
    "averageVolume10days", "averageDailyVolume10Day", "marketCap", "fiftyTwoWeekLow",
    "fiftyTwoWeekHigh", "priceToSalesTrailing12Months", "fiftyDayAverage",
    "quickRatio", "currentRatio", "totalRevenue", "debtToEquity", "revenuePerShare",
    "earningsGrowth", "revenueGrowth", "grossMargins", "ebitdaMargins", "operatingMargins",  "profitMargins",
    "previousClose", "open", "dayLow", "dayHigh", "regularMarketPreviousClose", "trailingPE", "forwardPE",
    "symbol", "shortName", "longName",  "industry", "sector", "currentPrice", "targetHighPrice", "targetLowPrice",
    "targetMeanPrice", "targetMedianPrice", "recommendationMean", "recommendationKey",
    "numberOfAnalystOpinions"
]

# Define a data type mapping for headers
data_type_map = {
    "industry": "STRING",
    "sector": "STRING",
    "fullTimeEmployees": "FLOAT",  # Integer field
    "auditRisk": "FLOAT",
    "boardRisk": "FLOAT",
    "compensationRisk": "FLOAT",
    "shareHolderRightsRisk": "FLOAT",
    "overallRisk": "FLOAT",
    "maxAge": "FLOAT",
    "priceHint": "FLOAT",
    "previousClose": "FLOAT",
    "open": "FLOAT",
    "dayLow": "FLOAT",
    "dayHigh": "FLOAT",
    "regularMarketPreviousClose": "FLOAT",
    "regularMarketOpen": "FLOAT",
    "regularMarketDayLow": "FLOAT",
    "regularMarketDayHigh": "FLOAT",
    "dividendRate": "FLOAT",
    "dividendYield": "FLOAT",
    "exDividendDate": "DATE",
    "payoutRatio": "FLOAT",
    "fiveYearAvgDividendYield": "FLOAT",
    "beta": "FLOAT",
    "trailingPE": "FLOAT",
    "forwardPE": "FLOAT",
    "volume": "FLOAT",
    "regularMarketVolume": "FLOAT",
    "averageVolume": "FLOAT",
    "averageVolume10days": "FLOAT",
    "averageDailyVolume10Day": "FLOAT",
    "marketCap": "FLOAT",
    "fiftyTwoWeekLow": "FLOAT",
    "fiftyTwoWeekHigh": "FLOAT",
    "priceToSalesTrailing12Months": "FLOAT",
    "fiftyDayAverage": "FLOAT",
    "twoHundredDayAverage": "FLOAT",
    "trailingAnnualDividendRate": "FLOAT",
    "trailingAnnualDividendYield": "FLOAT",
    "enterpriseValue": "FLOAT",
    "profitMargins": "FLOAT",
    "floatShares": "FLOAT",
    "sharesOutstanding": "FLOAT",
    "heldPercentInsiders": "FLOAT",
    "heldPercentInstitutions": "FLOAT",
    "impliedSharesOutstanding": "FLOAT",
    "bookValue": "FLOAT",
    "priceToBook": "FLOAT",
    "earningsQuarterlyGrowth": "FLOAT",
    "trailingEps": "FLOAT",
    "forwardEps": "FLOAT",
    "52WeekChange": "FLOAT",
    "lastDividendValue": "FLOAT",
    "lastDividendDate": "DATE",
    "exchange": "STRING",
    "quoteType": "STRING",
    "symbol": "STRING",
    "shortName": "STRING",
    "longName": "STRING",
    "currentPrice": "FLOAT",
    "targetHighPrice": "FLOAT",
    "targetLowPrice": "FLOAT",
    "targetMeanPrice": "FLOAT",
    "targetMedianPrice": "FLOAT",
    "recommendationMean": "FLOAT",
    "recommendationKey": "STRING",
    "numberOfAnalystOpinions": "FLOAT",
    "totalCash": "FLOAT",
    "totalCashPerShare": "FLOAT",
    "ebitda": "FLOAT",
    "totalDebt": "FLOAT",
    "quickRatio": "FLOAT",
    "currentRatio": "FLOAT",
    "totalRevenue": "FLOAT",
    "debtToEquity": "FLOAT",
    "revenuePerShare": "FLOAT",
    "returnOnAssets": "FLOAT",
    "returnOnEquity": "FLOAT",
    "freeCashflow": "FLOAT",
    "operatingCashflow": "FLOAT",
    "earningsGrowth": "FLOAT",
    "revenueGrowth": "FLOAT",
    "grossMargins": "FLOAT",
    "ebitdaMargins": "FLOAT",
    "operatingMargins": "FLOAT",
    "Today_Growth": "FLOAT",
    "Calculated_Score": "FLOAT",
    "Score_Recommendation": "STRING",
    "Conservative_Invs_Recom": "STRING",
    "Conservative_Invs_Reson": "STRING",
    "Growth_Invs_Recom": "STRING",
    "Growth_Invs_Reson": "STRING",
    "Momentum_Invs_Recom": "STRING",
    "Momentum_Invs_Reson": "STRING",
    "sector_rank": "INTEGER",
    "industry_rank": "INTEGER",
}

rank_headers = ["sector_rank", "industry_rank"]
ROW_COUNTER_FILE = os.path.join(MASTER_DIR, "nse_row_counter.txt")

# Initialize row_insert_order
def initialize_row_counter():
    if not os.path.exists(ROW_COUNTER_FILE):
        with open(ROW_COUNTER_FILE, "w") as f:
            f.write("1")  # Start counter at 0

def get_current_row_counter():
    with open(ROW_COUNTER_FILE, "r") as f:
        return int(f.read().strip())

def update_row_counter(new_value):
    with open(ROW_COUNTER_FILE, "w") as f:
        f.write(str(new_value))

# Initialize the row counter if not already done
initialize_row_counter()


# Add "Previous Day Date" to headers
# PREVIOUS_DAY_DATE = (ist_date - timedelta(days=1)).strftime('%Y-%m-%d') ist_now.strftime('%Y-%m-%d_%H-%M-%S')
PREVIOUS_DAY_DATETIME = ist_now.strftime('%Y-%m-%d %H:%M:%S')
#headers_with_date = ["row_insert_order", "PreviousDayDate", "Symbol_Input"] + headers

score_headers = ["Today_Growth", "Calculated_Score", "Score_Recommendation", "Conservative_Invs_Recom", "Conservative_Invs_Reson", "Growth_Invs_Recom", "Growth_Invs_Reson", "Momentum_Invs_Recom", "Momentum_Invs_Reson"]

def ensure_dataset_exists():
    try:
        bq_client.get_dataset(BQ_DATASET)
        log_message(f"Dataset '{BQ_DATASET}' exists.")
    except NotFound:
        dataset = bigquery.Dataset(f"{PROJECT_ID}.{BQ_DATASET}")
        bq_client.create_dataset(dataset)
        log_message(f"Created dataset '{BQ_DATASET}'.")

def ensure_table_exists():
    try:
        # Check if the table already exists
        table = bq_client.get_table(BQ_TABLE)
        log_message(f"Table '{BQ_TABLE}' already exists.")
    except NotFound:
        # Table does not exist, create it
        # Build the schema dynamically
        schema = [bigquery.SchemaField("row_insert_order", "INTEGER"), bigquery.SchemaField("PreviousDayDate", "DATETIME"), bigquery.SchemaField("Symbol_Input", "STRING"),] + [
                bigquery.SchemaField(header, data_type_map.get(header, "STRING"))
                for header in headers
                ] + [ bigquery.SchemaField(header, data_type_map.get(header, "STRING")) for header in score_headers ] +[ 
                bigquery.SchemaField(header, data_type_map.get(header, "STRING")) for header in rank_headers ]
        
        table = bigquery.Table(BQ_TABLE, schema=schema)
        bq_client.create_table(table)
        log_message(f"Created table '{BQ_TABLE}'.")
    except Exception as e:
        log_message(f"Error ensuring table exists: {e}")

def calculate_ranks(df, group_column, score_column, rank_column_name):
    """
    Calculate ranks within a group based on the score and append the rank as a new column.
    Handles NaN values in both group and score columns.
    """
    # Replace NaN values in score_column with a placeholder (-1) for rank calculation
    df[score_column] = pd.to_numeric(df[score_column], errors="coerce").fillna(-1)

    # Exclude rows where group_column is NaN
    valid_rows = ~df[group_column].isna()
    #print(df[group_column].isna())
    # Initialize the rank column with 0 as the default placeholder
    df[rank_column_name] = 0

    # Perform rank calculation only for valid rows
    df.loc[valid_rows, rank_column_name] = (
        df[valid_rows]
        .groupby(group_column)[score_column]
        .rank(ascending=False, method='dense', na_option='bottom')
        .fillna(0)
        .astype(int)
    )

    # Convert rank to integer (valid ranks) while keeping placeholder as 0
    df[rank_column_name] = df[rank_column_name].fillna(0).astype(int)
    return df

    
def append_to_csv(data_row, total_symbol):
    """Append a row of data to the CSV file, adding the header only if it's a new file."""
    write_header = not os.path.exists(MASTER_CSV_FILE_PATH)  # Check if file exists

    with open(MASTER_CSV_FILE_PATH, mode="a", newline="") as csv_file:
        writer = csv.writer(csv_file)
        if write_header:
            writer.writerow(["row_insert_order", "PreviousDayDate", "Symbol_Input"] + headers + score_headers)  # Add header row
            log_message(f"Header added to CSV file: {MASTER_CSV_FILE_PATH}")
        writer.writerow(data_row)
        log_message(f"Appended data to Master CSV file: {MASTER_CSV_FILE_PATH}")

        log_message(f" count: {processed_count}/{total_symbol}")
        
        # If it's the last row, calculate the ranks and update the file
        if processed_count==total_symbol:
            # Load the CSV file into DataFrame to calculate ranks
            df = pd.read_csv(MASTER_CSV_FILE_PATH)
            df.columns = df.columns.str.strip()
            # Ensure 'sector' and 'industry' columns exist, adjust accordingly to your file's structure
            if 'sector' in df.columns and 'industry' in df.columns and 'Calculated_Score' in df.columns:
                # Calculate ranks for sector and industry based on 'Calculated_Score' column
                log_message(f"Required columns found in Master CSV, starting rank calculation")

                # Handle NaN values in required columns
                df['Calculated_Score'] = df['Calculated_Score'].fillna(-1)  # Replace NaN in score_column with -1
                df['sector'] = df['sector'].fillna('Unknown')  # Replace NaN in sector with 'Unknown'
                df['industry'] = df['industry'].fillna('Unknown')  # Replace NaN in industry with 'Unknown'

                df = calculate_ranks(df, 'sector', 'Calculated_Score', 'sector_rank')
                df = calculate_ranks(df, 'industry', 'Calculated_Score', 'industry_rank')

                # Save the updated DataFrame back to the same CSV file, overwriting it
                df.to_csv(MASTER_CSV_FILE_PATH, index=False)
                log_message(f"Sector and Industry Rank calculation completed and saved to Master CSV file: {MASTER_CSV_FILE_PATH}")
            else:
                print(df.columns)
                
    """Append a row of data to the CSV file, adding the header only if it's a new file."""
    write_header = not os.path.exists(Daily_CSV_FILE_PATH)  # Check if file exists

    with open(Daily_CSV_FILE_PATH, mode="a", newline="") as csv_file:
        writer = csv.writer(csv_file)
        if write_header:
            writer.writerow(["row_insert_order", "PreviousDayDate", "Symbol_Input"] + headers + score_headers)  # Add header row
            log_message(f"Header added to CSV file: {Daily_CSV_FILE_PATH}")
        writer.writerow(data_row)
        log_message(f"Appended data to Daily CSV file: {Daily_CSV_FILE_PATH}")

        # If it's the last row, calculate the ranks and update the file
        if processed_count==total_symbol:
            # Load the CSV file into DataFrame to calculate ranks
            df = pd.read_csv(Daily_CSV_FILE_PATH)
            # Ensure 'sector' and 'industry' columns exist, adjust accordingly to your file's structure
            if 'sector' in df.columns and 'industry' in df.columns and 'Calculated_Score' in df.columns:
                log_message(f"Required columns found in Daily CSV, starting rank calculation")
                
                # Handle NaN values in required columns
                df['Calculated_Score'] = df['Calculated_Score'].fillna(-1)  # Replace NaN in score_column with -1
                df['sector'] = df['sector'].fillna('Unknown')  # Replace NaN in sector with 'Unknown'
                df['industry'] = df['industry'].fillna('Unknown')  # Replace NaN in industry with 'Unknown'
                
                # Calculate ranks for sector and industry based on 'Calculated_Score' column sector_rank industry_rank
                df = calculate_ranks(df, 'sector', 'Calculated_Score', 'sector_rank')
                df = calculate_ranks(df, 'industry', 'Calculated_Score', 'industry_rank')

                # Save the updated DataFrame back to the same CSV file, overwriting it
                df.to_csv(Daily_CSV_FILE_PATH, index=False)
                log_message(f"Sector and Industry Rank calculation completed and saved to Daily CSV file: {Daily_CSV_FILE_PATH}")
                
            # Load or create the Excel workbook
            if os.path.exists(EXCEL_FILE_PATH):
                workbook = load_workbook(EXCEL_FILE_PATH)
                log_message(f"Loaded existing Excel file. {EXCEL_FILE_PATH}")
            else:
                workbook = Workbook()
                workbook.remove(workbook.active)  # Remove default sheet
                log_message(f"Created new Excel file. {EXCEL_FILE_PATH}")
    
            # Check if sheet already exists, create if not
            sheet_name = f"NSE_{ist_date}"
            if sheet_name not in workbook.sheetnames:
                # Create a new sheet if it doesn't exist
                workbook.create_sheet(sheet_name)
                sheet = workbook[sheet_name]
                sheet.append(df.columns.tolist())  # Add headers
                log_message(f"New sheet created: {sheet_name}")
            else:
                sheet = workbook[sheet_name]
    
            # Append data to the sheet row by row
            for row in df.itertuples(index=False):
                sheet.append(row)
    
            # Freeze the first row and third column for better viewing
            sheet.freeze_panes = 'D2'  # Freeze everything above row 2 and to the left of column C
    
            # Save the updated Excel file
            workbook.save(EXCEL_FILE_PATH)
            log_message(f"Data successfully appended to Excel file: {EXCEL_FILE_PATH}_{sheet_name}")


def append_to_excel(data_row, total_symbol):
    """Append data to an Excel sheet, creating a new sheet for the day."""
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            workbook = load_workbook(EXCEL_FILE_PATH)
        else:
            workbook = Workbook()

        sheet_name = f"NSE_{ist_date}"
        if sheet_name not in workbook.sheetnames:
            # Create a new sheet if it doesn't exist
            workbook.create_sheet(sheet_name)
            sheet = workbook[sheet_name]
            # Add header row, including sector_rank and industry_rank
            sheet.append(
                ["row_insert_order", "PreviousDayDate", "Symbol_Input"] + headers + score_headers + ["sector_rank", "industry_rank"]
            )  # Ensure 'sector_rank' and 'industry_rank' headers are added
        else:
            sheet = workbook[sheet_name]

        # Append data row
        sheet.append(data_row)

        # If it's the last row, calculate the ranks and update the file
        if processed_count == total_symbol:
            log_message(f"Entered into rank calculation for Excel")

            # Fetch the column indices for 'Calculated_Score', 'sector', 'industry'
            calculated_score_index = score_headers.index('Calculated_Score') + 3  # Adjusted for offset (row_insert_order, PreviousDayDate, Symbol_Input)
            sector_index = headers.index('sector') + 3
            industry_index = headers.index('industry') + 3
            sector_rank_index = len(headers) + len(score_headers) + 4  # 'sector_rank' column index
            industry_rank_index = len(headers) + len(score_headers) + 5  # 'industry_rank' column index

            # Get all rows from the sheet to calculate ranks (skip header)
            rows = list(sheet.iter_rows(min_row=2, values_only=True))  # Skip header

            # Prepare data by replacing NaN/null values
            prepared_rows = []
            for row in rows:
                calculated_score = row[calculated_score_index] if row[calculated_score_index] is not None else -1  # Replace NaN in score with -1
                sector = row[sector_index] if row[sector_index] is not None else "Unknown"  # Replace NaN in sector with 'Unknown'
                industry = row[industry_index] if row[industry_index] is not None else "Unknown"  # Replace NaN in industry with 'Unknown'
                prepared_rows.append((row, calculated_score, sector, industry))

            # Group rows by sector and industry
            grouped_rows = defaultdict(list)
            for original_row, calculated_score, sector, industry in prepared_rows:
                grouped_rows[(sector, industry)].append((original_row, calculated_score))

            # Create lists to store ranks for sector and industry
            ranks_to_add = []  # Store rank values temporarily before writing them to the sheet

            # Assign ranks within each group based on the 'Calculated_Score'
            for (sector, industry), group_rows in grouped_rows.items():
                # Sort the rows by 'Calculated_Score' within the group (descending order)
                sorted_group = sorted(group_rows, key=lambda x: x[1], reverse=True)

                # Assign ranks within the group
                for rank, (original_row, _) in enumerate(sorted_group, start=1):
                    ranks_to_add.append((original_row, rank, rank))  # sector_rank and industry_rank

            # Write ranks to the corresponding columns in the sheet
            for row_idx, (original_row, sector_rank, industry_rank) in enumerate(ranks_to_add, start=2):  # Start from row 2
                sheet.cell(row=row_idx, column=sector_rank_index, value=sector_rank)  # Add sector rank
                sheet.cell(row=row_idx, column=industry_rank_index, value=industry_rank)  # Add industry rank

            log_message(f"Rank calculation completed and appended to Excel file: {EXCEL_FILE_PATH}")

        # Freeze the first row and third column for better viewing
        sheet.freeze_panes = 'D2'  # Freeze everything above row 2 and to the left of column C

        # Save the workbook after appending the data and ranks
        workbook.save(EXCEL_FILE_PATH)
        log_message(f"Data appended to Excel file: {EXCEL_FILE_PATH}")

    except Exception as e:
        log_message(f"Error saving to Excel: {e}")


def validate_input(value, min_val=None):
    if value is None or pd.isna(value) or not isinstance(value, (int, float)):
        return None
    if min_val is not None and value < min_val:
        return None
    return value
        
def calculate_individual_scores(pe, dividend_yield, earnings_growth):

    dividend_yield = dividend_yield * 100

    # P/E Ratio Scoring Logic (Lower is better)
    if pe is not None:
        if pe <= 10:
            pe_score = 5
        elif pe <= 20:
            pe_score = 4
        elif pe <= 30:
            pe_score = 3
        elif pe <= 50:
            pe_score = 2
        else:
            pe_score = 1
    else:
        pe_score = 0

    # Dividend Yield Scoring Logic (Higher is better)
    if dividend_yield is not None:
        if dividend_yield > 4:
            dividend_score = 5
        elif dividend_yield > 3:
            dividend_score = 4
        elif dividend_yield > 2:
            dividend_score = 3
        elif dividend_yield > 1:
            dividend_score = 2
        else:
            dividend_score = 1
    else:
        dividend_score = 0

    # Earnings Growth Scoring Logic (Higher is better)
    if earnings_growth is not None:
        if earnings_growth > 20:
            earnings_growth_score = 5
        elif earnings_growth > 10:
            earnings_growth_score = 4
        elif earnings_growth > 5:
            earnings_growth_score = 3
        elif earnings_growth > 0:
            earnings_growth_score = 2
        else:
            earnings_growth_score = 1
    else:
        earnings_growth_score = 0

    # If any of the scores are invalid, return "None"
    if pe_score == 0 or dividend_score == 0 or earnings_growth_score == 0:
        return None

    # Weighted total score (scaled to 1-5)
    total_score = (pe_score * 0.4) + (dividend_score * 0.3) + (earnings_growth_score * 0.3)
    total_score = round(total_score, 1)
    
    # Assign Calculated Recommendation
    if total_score <= 1.5:
        recommendation = "Strong Buy"
    elif total_score <= 2.5:
        recommendation = "Buy"
    elif total_score <= 3.5:
        recommendation = "Hold"
    elif total_score <= 4.5:
        recommendation = "Underperform"
    else:
        recommendation = "Sell"

    return total_score, recommendation
    #return round(total_score, 1)

def analyze_stock_with_profiles(info):
    recommendations = []
    
    try:
        # Extract relevant fields
        beta = info.get('beta', 'N/A')
        pe_ratio = info.get('trailingPE', 'N/A')
        forward_pe = info.get('forwardPE', 'N/A')
        dividend_yield = info.get('dividendYield', 'N/A')
        price_to_book = info.get('priceToBook', 'N/A')
        profit_margins = info.get('profitMargins', 'N/A')
        revenue_growth = info.get('revenueGrowth', 'N/A')
        high_52w = info.get('fiftyTwoWeekHigh', 'N/A')
        low_52w = info.get('fiftyTwoWeekLow', 'N/A')
        recommendation_mean = info.get('recommendationMean', 'N/A')
        current_price = info.get('currentPrice', 'N/A')

        # 1. Conservative Investor (Low Risk, Income-Focused)
        if beta != 'N/A' and beta < 1:
            conservative_reason = "Low Beta (less volatile than the market)"
        else:
            conservative_reason = "High Beta (more volatile)"
        
        if dividend_yield != 'N/A' and dividend_yield > 0.03:
            conservative_reason += "- Pays a good dividend (>3%)"
        
        if price_to_book != 'N/A' and price_to_book < 1:
            conservative_reason += "- Price-to-Book ratio (<1) indicates undervalued assets"
        elif price_to_book != 'N/A' and price_to_book < 2:
            conservative_reason += "- Price-to-Book ratio (<2) indicates potential for growth"
        
        recommendations.append({
            "Cal_Investment_Profile": "Conservative Investor",
            "Cal_Recommendation": "Buy" if dividend_yield != 'N/A' and dividend_yield > 0.03 else "Hold",
            "Cal_Reason": conservative_reason
        })

        # 2. Growth Investor (Focus on High Growth)
        growth_reason = []
        if forward_pe != 'N/A' and forward_pe < 20:
            growth_reason.append("Low Forward P/E (<20) indicates growth potential")
        if revenue_growth != 'N/A' and revenue_growth > 0.1:
            growth_reason.append("Strong Revenue Growth (>10%)")
        if profit_margins != 'N/A' and profit_margins > 0.2:
            growth_reason.append("Highly Profitable with margins > 20%")
        
        if growth_reason:
            recommendations.append({
                "Cal_Investment_Profile": "Growth Investor",
                "Cal_Recommendation": "Buy" if forward_pe != 'N/A' and forward_pe < 20 else "Hold",
                "Cal_Reason": "- ".join(growth_reason)
            })
        else:
            # Add a default entry with None as the reason
            recommendations.append({
                "Cal_Investment_Profile": "Growth Investor",
                "Cal_Recommendation": "None",
                "Cal_Reason": "None"
            })

        # 3. Momentum Investor (Focus on Recent Trends)
        momentum_reason = []
        if high_52w != 'N/A' and low_52w != 'N/A':
            price_position = (current_price - low_52w) / (high_52w - low_52w)
            if price_position > 0.75:
                momentum_reason.append("Trading near its 52-week high (bullish momentum)")
            elif price_position < 0.25:
                momentum_reason.append("Trading near its 52-week low (bearish momentum)")
        else:
            momentum_reason.append("None")
       
        recommendations.append({
            "Cal_Investment_Profile": "Momentum Investor",
            "Cal_Recommendation": "Buy" if price_position > 0.75 else "Hold",
            "Cal_Reason": "- ".join(momentum_reason) if momentum_reason else "No strong momentum"
        })

    except Exception as e:
        recommendations.append({
            "Cal_Investment_Profile": "Error",
            "Cal_Recommendation": "None",
            "Cal_Reason": "None"
        })
    
    return recommendations
    
    
def fetch_and_update_stock_data(symbol, total_symbol):
    try:
        # Read the current row counter
        current_counter = get_current_row_counter()

        log_message(f"Life count: {current_counter} Fetching data for {symbol} ...")
        stock = yf.Ticker(symbol)
        info = stock.info
        
        # Safely access data with default values
        pe_ratio = info.get('trailingPE', 0)
        dividend_yield = info.get('dividendYield', 0)
        earnings_growth = info.get('earningsQuarterlyGrowth', 0)        

        # Calculate score (assuming the `calculate_individual_scores` function is defined)
        score, score_recommendation = calculate_individual_scores(pe_ratio, dividend_yield, earnings_growth)

        cal_recom = analyze_stock_with_profiles(info)

        current_price = info.get('currentPrice', 'N/A')
        previous_close = info.get('previousClose', 'N/A')

        # Calculate today's growth percentage
        if current_price != 'N/A' and previous_close != 'N/A' and previous_close != 0:
            today_growth_percentage = ((current_price - previous_close) / previous_close) * 100
            today_growth_percentage = round(today_growth_percentage, 2)
        else:
            today_growth_percentage = 'N/A'

        PREVIOUS_DAY_DATETIME = ist_now.strftime('%Y-%m-%d %H:%M:%S')
        # Extract data and include the Previous Day Date
        info_row = [current_counter, PREVIOUS_DAY_DATETIME, symbol] + [info.get(key, '') for key in headers]  + [today_growth_percentage, score, score_recommendation]

        for recom in cal_recom:
            #print(recom)
            #info_row.append(recom.get("Cal_Investment_Profile", ""))
            info_row.append(recom.get("Cal_Recommendation", ""))
            info_row.append(recom.get("Cal_Reason", ""))
    
        # Increment row_insert_order for the next row
        current_counter += 1
        update_row_counter(current_counter)
        
        # Append data to CSV and Excel
        append_to_csv(info_row, total_symbol)
        #append_to_excel(info_row, total_symbol)
       
        return info_row
    except Exception as e:
        log_message(f"Error fetching data for {symbol}: {e}")
        return None

# Add process data
def preprocess_data(csv_file_path):
    """
    Preprocess the CSV file to ensure data types are correct based on the BigQuery schema.
    If incorrect types are detected, log the error and attempt to fix them.
    """

    processed_rows = []
    errors = []

    try:
        with open(csv_file_path, "r") as csv_file:
            reader = csv.DictReader(csv_file)
            for row in reader:
                processed_row = {}
                for key, value in row.items():
                    expected_type = data_type_map.get(key, "STRING")
                    try:
                        if expected_type == "STRING":
                            processed_row[key] = value.strip() if value else ""  # Handle empty strings as None
                        elif expected_type == "INTEGER":
                            processed_row[key] = int(value) if value else 0
                        elif expected_type == "FLOAT":
                            try:
                                # Attempt to convert the value to a float
                                processed_row[key] = float(value)
                            except (ValueError, TypeError):
                                # If conversion fails, set it to None
                                processed_row[key] = 0
                        elif expected_type == "DATETIME":
                            processed_row[key] = (
                                datetime.strptime(value, "%Y-%m-%d %H:%M:%S") 
                                if value 
                                else datetime(1990, 1, 1, 0, 0, 0)  # Default to '1990-01-01 00:00:00' if no value is provided
                            )
                        elif expected_type == "DATE":
                            try:
                                # Check if value is a Unix timestamp and convert it
                                if value.isdigit():
                                    processed_row[key] = datetime.fromtimestamp(int(value)).date()
                                else:
                                    # Parse date string in the format "YYYY-MM-DD"
                                    processed_row[key] = datetime.strptime(value, "%Y-%m-%d").date()
                            except Exception:
                                # Handle invalid or missing date values with a default date
                                processed_row[key] = datetime(1990, 1, 1).date()
                        else:  # STRING
                            processed_row[key] = ""
                    except (ValueError, TypeError, KeyError) as ve:
                        errors.append(
                            f"Row {processed_count}, Field '{key}' with value '{value}' failed conversion to {expected_type}: {ve}"
                        )
                        processed_row[key] = ""  # Default to None on error
                        
                    # Validate that the processed row has consistent column counts
                    #if len(processed_row) == len(row):
                    #    processed_rows.append(processed_row)
                    #else:
                    #    errors.append(f"Row {row_num} has inconsistent column counts.")
                    
                processed_rows.append(processed_row)
    except Exception as e:
        log_message(f"Error reading or processing CSV file: {e}")
    
    # Log errors, if any
    if errors:
        log_message(f"Data type errors detected during preprocessing:\n" + "\n".join(errors))
    
    return processed_rows

def load_data_to_bigquery():
    """Load data from the preprocessed CSV file into BigQuery."""
    try:
        processed_data = preprocess_data(Daily_CSV_FILE_PATH)
        
        # Write processed data back to a temporary CSV for BigQuery loading
        temp_csv_path = "temp_processed.csv"
        
        # Check if the file exists, and delete it if it does
        if os.path.exists(temp_csv_path):
            os.remove(temp_csv_path)
            log_message(f"Deleted the file before start - {temp_csv_path}.")
            
        with open(temp_csv_path, "w", newline="") as temp_csv:
            writer = csv.DictWriter(temp_csv, fieldnames=processed_data[0].keys())
            writer.writeheader()  # Write headers
            writer.writerows(processed_data)  # Write processed rows
        
        log_message(f"Start to load data to BigQuery from {temp_csv_path}.")

        # Load the processed data into BigQuery
        with open(temp_csv_path, "rb") as csv_file:
            job_config = bigquery.LoadJobConfig(
                source_format=bigquery.SourceFormat.CSV,
                skip_leading_rows=1,  # Skip header row
                write_disposition="WRITE_APPEND",  # Append data schema=schema
                autodetect=False,
                max_bad_records=5000,  # Tolerate up to 50 bad rows
               # ignore_unknown_values=True,  # Ignore unexpected columns
            )
            load_job = bq_client.load_table_from_file(
                csv_file, BQ_TABLE, job_config=job_config
            )
            load_job.result()  # Wait for the job to complete
            log_message("Job Done")
            # Check for errors
            if load_job.errors:
                log_message(f"Errors encountered during loading: {load_job.errors}")
            else:
                log_message("Data loaded successfully, no errors.")
            log_message(f"Data successfully loaded to BigQuery from {temp_csv_path}.")
    except Exception as e:
        log_message(f"Error loading data to BigQuery: {e}")

# Process each symbol
processed_count = 0

# Process each symbol
for symbol in symbols:
    processed_count += 1
    fetch_and_update_stock_data(symbol, len(symbols))

    # Add a delay to avoid rate-limiting
    time.sleep(0.7)
    log_message(f"Processed {processed_count}/{len(symbols)} symbols.")


# Define BigQuery dataset and table with the project ID
PROJECT_ID = "stockautomation-442015"  # Replace with your project ID
BQ_DATASET = "nse_stock_score_master"  # Replace with your dataset name
BQ_TABLE = f"{PROJECT_ID}.{BQ_DATASET}.daily_nse_stock_master"  # Fully-qualified table name

# BigQuery authentication
bq_client = bigquery.Client.from_service_account_json(SERVICE_ACCOUNT_FILE)

# Ensure dataset and table exist in BigQuery
ensure_dataset_exists()
ensure_table_exists()

# Load the data into BigQuery from the CSV file
load_data_to_bigquery()

log_message("Script execution completed.")
